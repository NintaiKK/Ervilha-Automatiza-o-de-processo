import sqlite3
from openpyxl import load_workbook
import logging

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%H:%M:%S'
)

def parse_fonte_estruturado():
    """
    Parser específico para a estrutura da planilha fonte.xlsx
    com colunas: NOME, ENDERECO, CONTATO, DEMANDA, INFORMACOES, ENCAMINHAMENTO
    """
    try:
        # Carregar a planilha
        workbook = load_workbook('fonte.xlsx')
        sheet = workbook.active
        
        # Conectar ao SQLite
        conn = sqlite3.connect('fonte_estruturado.db')
        cursor = conn.cursor()
        
        # Criar tabela com estrutura específica
        create_table_sql = """
        CREATE TABLE IF NOT EXISTS demandas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT,
            endereco TEXT,
            contato TEXT,
            demanda TEXT,
            informacoes TEXT,
            encaminhamento TEXT,
            data_importacao TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """
        cursor.execute(create_table_sql)
        
        # SQL para inserção
        insert_sql = """
        INSERT INTO demandas (nome, endereco, contato, demanda, informacoes, encaminhamento)
        VALUES (?, ?, ?, ?, ?, ?)
        """
        
        # Processar linhas (assumindo que a primeira linha é cabeçalho)
        linhas_processadas = 0
        linhas_com_erro = 0
        
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
            try:
                # Extrair dados baseado na estrutura esperada
                # Ajuste os índices conforme a ordem real das colunas na sua planilha
                nome = str(row[0]) if row[0] is not None else ""
                endereco = str(row[1]) if row[1] is not None else ""
                contato = str(row[2]) if row[2] is not None else ""
                demanda = str(row[3]) if row[3] is not None else ""
                informacoes = str(row[4]) if row[4] is not None else ""
                encaminhamento = str(row[5]) if row[5] is not None else ""
                
                # Verificar se a linha não está vazia
                if any([nome, endereco, contato, demanda, informacoes, encaminhamento]):
                    cursor.execute(insert_sql, (
                        nome.strip(),
                        endereco.strip(),
                        contato.strip(),
                        demanda.strip(),
                        informacoes.strip(),
                        encaminhamento.strip()
                    ))
                    linhas_processadas += 1
                    
                    # Log a cada 50 linhas processadas
                    if linhas_processadas % 50 == 0:
                        logging.info(f"Processadas {linhas_processadas} linhas...")
                        
            except Exception as e:
                logging.warning(f"Erro na linha {row_num}: {e}")
                linhas_com_erro += 1
                continue
        
        # Commit e fechar
        conn.commit()
        conn.close()
        workbook.close()
        
        logging.info(f"✅ Importação concluída: {linhas_processadas} linhas processadas, {linhas_com_erro} erros")
        
        return linhas_processadas, linhas_com_erro
        
    except FileNotFoundError:
        logging.error("❌ Arquivo fonte.xlsx não encontrado!")
        return 0, 0
    except Exception as e:
        logging.error(f"❌ Erro geral: {e}")
        return 0, 0

def analisar_estrutura_planilha():
    """
    Função para analisar a estrutura real da planilha
    """
    try:
        workbook = load_workbook('fonte.xlsx')
        sheet = workbook.active
        
        print("\n=== ANÁLISE DA PLANILHA ===")
        print(f"Total de linhas: {sheet.max_row}")
        print(f"Total de colunas: {sheet.max_column}")
        
        # Mostrar cabeçalhos
        headers = [cell.value for cell in sheet[1]]
        print(f"\nCabeçalhos encontrados ({len(headers)} colunas):")
        for i, header in enumerate(headers):
            print(f"  Coluna {i+1}: '{header}'")
        
        # Mostrar algumas linhas de exemplo
        print(f"\nPrimeiras 3 linhas de dados:")
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, max_row=4, values_only=True), 2):
            print(f"  Linha {row_num}: {row}")
        
        workbook.close()
        
    except Exception as e:
        print(f"Erro na análise: {e}")

def consultar_dados_importados():
    """
    Consulta os dados importados com estatísticas
    """
    try:
        conn = sqlite3.connect('fonte_estruturado.db')
        cursor = conn.cursor()
        
        print("\n=== DADOS IMPORTADOS ===")
        
        # Estatísticas básicas
        cursor.execute("SELECT COUNT(*) FROM demandas")
        total = cursor.fetchone()[0]
        print(f"Total de registros: {total}")
        
        # Contar por tipo de demanda (exemplo)
        cursor.execute("""
            SELECT demanda, COUNT(*) as quantidade 
            FROM demandas 
            WHERE demanda != '' 
            GROUP BY demanda 
            ORDER BY quantidade DESC
            LIMIT 10
        """)
        print(f"\nTop 10 demandas:")
        for demanda, quant in cursor.fetchall():
            print(f"  {demanda}: {quant}")
        
        # Mostrar alguns registros
        cursor.execute("""
            SELECT nome, contato, demanda, encaminhamento 
            FROM demandas 
            LIMIT 5
        """)
        print(f"\nExemplo de registros:")
        for nome, contato, demanda, encaminhamento in cursor.fetchall():
            print(f"  Nome: {nome}")
            print(f"  Contato: {contato}")
            print(f"  Demanda: {demanda}")
            print(f"  Encaminhamento: {encaminhamento}")
            print("  ---")
        
        conn.close()
        
    except Exception as e:
        print(f"Erro na consulta: {e}")

# Versão com mapeamento flexível de colunas
def parse_fonte_flexivel(mapeamento_colunas=None):
    """
    Versão flexível que permite mapear colunas personalizadas
    """
    if mapeamento_colunas is None:
        # Mapeamento padrão baseado na descrição
        mapeamento_colunas = {
            'nome': 0,
            'endereco': 1, 
            'contato': 2,
            'demanda': 3,
            'informacoes': 4,
            'encaminhamento': 5
        }
    
    try:
        workbook = load_workbook('fonte.xlsx')
        sheet = workbook.active
        conn = sqlite3.connect('fonte_flexivel.db')
        cursor = conn.cursor()
        
        # Criar tabela
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS demandas_flex (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nome TEXT, endereco TEXT, contato TEXT,
                demanda TEXT, informacoes TEXT, encaminhamento TEXT
            )
        """)
        
        insert_sql = """
            INSERT INTO demandas_flex (nome, endereco, contato, demanda, informacoes, encaminhamento)
            VALUES (?, ?, ?, ?, ?, ?)
        """
        
        linhas_importadas = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                # Extrair dados usando mapeamento
                dados = {}
                for campo, indice in mapeamento_colunas.items():
                    if indice < len(row):
                        valor = row[indice]
                        dados[campo] = str(valor).strip() if valor is not None else ""
                    else:
                        dados[campo] = ""
                
                # Inserir se houver dados
                if any(dados.values()):
                    cursor.execute(insert_sql, (
                        dados['nome'], dados['endereco'], dados['contato'],
                        dados['demanda'], dados['informacoes'], dados['encaminhamento']
                    ))
                    linhas_importadas += 1
                    
            except Exception as e:
                continue
        
        conn.commit()
        conn.close()
        workbook.close()
        
        print(f"✅ Importação flexível: {linhas_importadas} linhas")
        
    except Exception as e:
        print(f"❌ Erro: {e}")

if __name__ == "__main__":
    print("🚀 Iniciando processamento da planilha fonte.xlsx...")
    
    # Primeiro, analisar a estrutura
    analisar_estrutura_planilha()
    
    # Importar dados
    sucesso, erros = parse_fonte_estruturado()
    
    # Mostrar resultados
    if sucesso > 0:
        consultar_dados_importados()
        
        print(f"\n📊 RESUMO FINAL:")
        print(f"   ✅ Linhas importadas: {sucesso}")
        print(f"   ❌ Erros: {erros}")
        print(f"   💾 Banco: fonte_estruturado.db")
        print(f"   📋 Tabela: demandas")
    else:
        print("❌ Nenhum dado foi importado!")